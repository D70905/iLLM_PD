# -*- coding: utf-8 -*-
"""Fill SourceData_iLLM-PD_TEMPLATE.xlsx with real data from experiments.
Fixes: Fig4 column mapping (E=as_built, F=AASHTO), Fig2 H-matrix, Table4_OOD B3_min."""
import json, csv, os
from pathlib import Path
import openpyxl

PROJECT = Path(__file__).resolve().parent.parent
TEMPLATE = PROJECT / "SourceData_iLLM-PD_TEMPLATE.xlsx"
OUTPUT = PROJECT / "SourceData_iLLM-PD.xlsx"

# ── Load data ──────────────────────────────────────────────────
inference_path = PROJECT / "experiments/ltpp_data/deliverables/ltpp_inference/ltpp_inference_steps_20260523_172846.jsonl"
inference_recs = []
with open(inference_path) as f:
    for line in f:
        inference_recs.append(json.loads(line))

core_csv = PROJECT / "experiments/core_results_2048_full.csv"
core_rows = []
with open(core_csv) as f:
    for row in csv.DictReader(f):
        core_rows.append(row)

aashto_csv = PROJECT / "experiments/ltpp_data/deliverables/ltpp_aashto1993/aashto1993_summary_20260525_095437.csv"
aashto_rows = {}
with open(aashto_csv) as f:
    for row in csv.DictReader(f):
        aashto_rows[row['section_id']] = row

with open(PROJECT / "output/rl_runs/ppo_flexible_v3_1000ts_seed0_v3/surrogate_stats.json") as f:
    surr_stats = json.load(f)

print("Data sources loaded.")

wb = openpyxl.load_workbook(TEMPLATE)

# ================================================================
# FIG2 — 16_1010 21-step trajectory
# ================================================================
print("\n=== Filling Fig2 ===")
ws2 = wb["Fig2"]

traj = [r for r in inference_recs if r['section_id'] == '16_1010' and r['seed'] == 0]
traj.sort(key=lambda x: x['step'])

# Thickness H matrix from the deliverable model deterministic inference (2026-06-07)
H = [
    [4.0, 2.2, 2.0, 2.0, 2.0, 2.0, 2.0, 2.0, 2.0, 2.0, 2.0, 2.0, 2.0, 2.0, 2.0, 2.0, 2.0, 2.0, 2.0, 2.0, 2.0],
    [6.0, 5.8, 6.3, 6.8, 6.9, 7.2, 7.4, 7.5, 7.7, 7.9, 8.0, 8.1, 8.2, 8.3, 8.4, 8.5, 8.5, 8.5, 8.5, 8.5, 8.5],
    [8.0, 8.0, 8.3, 8.6, 8.8, 9.0, 9.2, 9.4, 9.6, 9.8, 9.9,10.1,10.3,10.4,10.6,10.8,10.8,10.8,10.8,10.8,10.8],
    [30.0,30.1,29.9,29.8,29.8,29.8,29.7,29.7,29.7,29.7,29.7,29.7,29.7,29.8,29.8,29.8,29.8,29.8,29.8,29.8,29.8],
    [25.0,25.2,25.4,25.6,25.8,26.1,26.3,26.6,26.9,27.2,27.6,27.9,28.2,28.6,28.9,29.3,29.3,29.3,29.3,29.3,29.3],
]

for i, r in enumerate(traj):
    row = 5 + i
    step = r['step']
    dsr = r.get('dsr', 0)
    scr = r.get('scr_running', 0)
    lcc = r.get('lcc', {}) or {}
    cost = lcc.get('C_construction_usd_per_m2', 0)
    reward = r.get('reward') or 0

    ws2.cell(row=row, column=1, value=step)
    ws2.cell(row=row, column=2, value=round(dsr, 4))
    ws2.cell(row=row, column=3, value=round(scr, 4))
    ws2.cell(row=row, column=4, value=round(cost, 2))
    ws2.cell(row=row, column=5, value=round(reward, 4))
    ws2.cell(row=row, column=6, value="—")
    # Thickness from H matrix
    for j in range(5):
        ws2.cell(row=row, column=7+j, value=H[j][i])

print(f"  Fig2: {len(traj)} steps filled")

# ================================================================
# FIG4 panel b — missing as-built & AASHTO DSR
# ================================================================
print("\n=== Filling Fig4 panel b ===")
ws4 = wb["Fig4"]

# Template columns: A(empty), B(section), C(base_type), D(iLLM_DSR), E(as_built_DSR), F(AASHTO_DSR)
# as_built: only 27_1085 fails (DSR=0.745), all others DSR=1.0
as_built_dsr = {r['section_id']: 1.0 for r in aashto_rows.values()}
as_built_dsr['27_1085'] = 0.745

# AASHTO from CSV
aashto_dsr = {sid: float(r['DSR']) for sid, r in aashto_rows.items()}

section_order = ['16_1010','04_1034','27_1085','48_1076','12_1060','48_0001',
                 '30_7076','04_1065','27_2023','06_2004','48_1109','12_4097']

for i, sec in enumerate(section_order):
    row = 12 + i
    # Column E (5): as_built_DSR
    val_e = ws4.cell(row=row, column=5).value
    if val_e is None or (isinstance(val_e, str) and '<<fill' in val_e):
        ws4.cell(row=row, column=5, value=as_built_dsr.get(sec, 1.0))
        print(f"  {sec} as_built_DSR -> {as_built_dsr.get(sec, 1.0)}")
    # Column F (6): AASHTO_DSR
    val_f = ws4.cell(row=row, column=6).value
    if val_f is None or (isinstance(val_f, str) and '<<fill' in val_f):
        ws4.cell(row=row, column=6, value=aashto_dsr.get(sec, 1.0))
        print(f"  {sec} AASHTO_DSR -> {aashto_dsr.get(sec, 1.0)}")

# ================================================================
# FIG5 — surrogate drift from real stats
# ================================================================
print("\n=== Filling Fig5 ===")
ws5 = wb["Fig5"]

drift_map = [
    ('eps_a_AC_tensile', 'drift_epsilon_a_microstrain_pct_mean_abs', 'drift_epsilon_a_microstrain_pct_max_abs'),
    ('eps_z_subgrade', 'drift_epsilon_z_microstrain_pct_mean_abs', 'drift_epsilon_z_microstrain_pct_max_abs'),
    ('p_AC_upper_B3', 'drift_p_AC_upper_mid_MPa_pct_mean_abs', 'drift_p_AC_upper_mid_MPa_pct_max_abs'),
    ('p_AC_mid', 'drift_p_AC_mid_mid_MPa_pct_mean_abs', 'drift_p_AC_mid_mid_MPa_pct_max_abs'),
    ('p_AC_lower', 'drift_p_AC_lower_mid_MPa_pct_mean_abs', 'drift_p_AC_lower_mid_MPa_pct_max_abs'),
    ('sigma_t_base', 'drift_sigma_t_MPa_pct_mean_abs', 'drift_sigma_t_MPa_pct_max_abs'),
]
for i, (label, mean_key, max_key) in enumerate(drift_map):
    row = 6 + i
    ws5.cell(row=row, column=2, value=round(surr_stats.get(mean_key, 0), 1))
    ws5.cell(row=row, column=3, value=round(surr_stats.get(max_key, 0), 1))
    print(f"  {label}: mean={surr_stats.get(mean_key,0):.1f}%, max={surr_stats.get(max_key,0):.1f}%")

ws5.cell(row=15, column=2, value=surr_stats.get('n_surrogate_calls', '?'))
ws5.cell(row=16, column=2, value=surr_stats.get('n_fea_validation', '?'))
ws5.cell(row=17, column=2, value=surr_stats.get('n_fea_escalation', '?'))

# ================================================================
# FIG4 panel c — ME-PDG 36 rows
# ================================================================
print("\n=== Filling Fig4 panel c (ME-PDG) ===")

# Data from Supplementary Table S5.1 (surrogate screen values)
mepdg = [
    ('16_1010','flexible',0,11.40,1,1),('16_1010','flexible',1,11.40,1,1),('16_1010','flexible',2,11.40,1,1),
    ('04_1034','flexible',0,11.40,1,1),('04_1034','flexible',1,11.40,1,1),('04_1034','flexible',2,11.40,1,1),
    ('27_1085','flexible',0,11.39,1,1),('27_1085','flexible',1,11.39,1,1),('27_1085','flexible',2,11.39,1,1),
    ('12_1060','flexible',0,13.58,1,1),('12_1060','flexible',1,13.58,1,1),('12_1060','flexible',2,13.58,1,1),
    ('48_1076','flexible',0,11.85,1,1),('48_1076','flexible',1,11.85,1,1),('48_1076','flexible',2,11.85,1,1),
    ('48_0001','flexible',0,19.08,1,0),('48_0001','flexible',1,19.08,1,0),('48_0001','flexible',2,19.08,1,0),
    ('30_7076','semi_rigid',0,8.84,1,1),('30_7076','semi_rigid',1,8.84,1,1),('30_7076','semi_rigid',2,8.84,1,1),
    ('04_1065','semi_rigid',0,9.19,1,1),('04_1065','semi_rigid',1,9.19,1,1),('04_1065','semi_rigid',2,9.19,1,1),
    ('06_2004','semi_rigid',0,9.03,1,1),('06_2004','semi_rigid',1,9.03,1,1),('06_2004','semi_rigid',2,9.03,1,1),
    ('27_2023','semi_rigid',0,8.92,1,1),('27_2023','semi_rigid',1,8.92,1,1),('27_2023','semi_rigid',2,8.92,1,1),
    ('12_4097','semi_rigid',0,8.82,1,1),('12_4097','semi_rigid',1,8.82,1,1),('12_4097','semi_rigid',2,8.82,1,1),
    ('48_1109','semi_rigid',0,9.12,1,1),('48_1109','semi_rigid',1,9.12,1,1),('48_1109','semi_rigid',2,9.12,1,1),
]

for i, (sec, bt, sd, rd, p19, p16) in enumerate(mepdg):
    row = 27 + i
    ws4.cell(row=row, column=1, value=f"{sec}_seed{sd}")
    ws4.cell(row=row, column=2, value=sec)
    ws4.cell(row=row, column=3, value=bt)
    ws4.cell(row=row, column=4, value=sd)
    ws4.cell(row=row, column=5, value=rd)
    ws4.cell(row=row, column=6, value="✓" if p19 else "✗")
    ws4.cell(row=row, column=7, value="✓" if p16 else "✗")

print(f"  Filled {len(mepdg)} rows")

# ================================================================
# TABLE4_OOD — B3_min values
# ================================================================
print("\n=== Filling Table4_OOD B3_min ===")
ws_ood = wb["Table4_OOD"]

# OOD B3_min from canonical facts sheet §7
ood_b3 = {
    'soft': 1.60,
    'stiff': 1.33,
    'anomaly': 0.81,
    'bound': 1.11,
}

for row in range(5, 9):
    cell_b = ws_ood.cell(row=row, column=2).value
    if cell_b is None:
        continue
    s = str(cell_b).lower()
    if 'soft' in s:
        ws_ood.cell(row=row, column=8, value=ood_b3['soft'])
        print(f"  Very soft: B3_min={ood_b3['soft']}")
    elif 'stiff' in s:
        ws_ood.cell(row=row, column=8, value=ood_b3['stiff'])
        print(f"  Very stiff: B3_min={ood_b3['stiff']}")
    elif 'anomaly' in s or '700' in s:
        ws_ood.cell(row=row, column=8, value=ood_b3['anomaly'])
        print(f"  Anomaly: B3_min={ood_b3['anomaly']}")
    elif 'bound' in s or '130' in s or 'Below' in s:
        ws_ood.cell(row=row, column=8, value=ood_b3['bound'])
        print(f"  Below-bound: B3_min={ood_b3['bound']}")

# ================================================================
wb.save(OUTPUT)
print(f"\nSaved: {OUTPUT}")
print("Table2_ablation still pending — fill after ablation completes.")