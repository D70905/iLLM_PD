# -*- coding: utf-8 -*-
"""
mc_estar.py — 从 NCAT 动态模量主曲线 Excel 计算任意温度/频率下的 |E*|。

ME-PDG 车辙模型按"逐温度档"累积损伤, 需要在多个路面温度下取沥青层模量;
build_ncat_case.py 只取了 20°C/10Hz 一个点, 不够。本模块用每个 sheet 里已拟合
好的 sigmoid 主曲线系数(Max/Min/Beta/Gamma)+ Arrhenius 频移(活化能 EA, 参考温
度 20°C)给出 E*(T, f)。公式已对照 sheet 内置的拟合值逐点验证(见 __main__ 自检)。

sigmoid:  log10|E*|(ksi) = logMin + (logMax-logMin)/(1+exp(Beta + Gamma*log10(wr)))
频移:     log10 aT = (EA / (2.303*R)) * (1/T_K - 1/Tref_K),  R=8.314, Tref=20°C
          wr = f * aT     (参考温度处 aT=1)
"""
import math, re
import openpyxl

KSI_TO_MPA = 6.894757
R_GAS = 8.314
LOG10 = 2.302585
TREF_C_DEFAULT = 20.0

# sheet 名映射: 段 -> (面层 sheet, 共用基层 sheet)。aging ∈ {"RH","CA"}。
def sheet_names(section: str, aging: str = "RH"):
    surf = f"{section}-1 {aging}"
    base = f"CG Base {aging}"
    return surf, base


def _find_coeffs(ws):
    """定位 'Max E* (ksi) | Min E* | Beta | Gamma | EA' 标题行, 读下一行的数值。"""
    for r in range(1, min(ws.max_row, 130) + 1):
        c0 = ws.cell(r, 1).value
        if isinstance(c0, str) and "Max E*" in c0:
            vals = [ws.cell(r + 1, c).value for c in range(1, 6)]  # Max,Min,Beta,Gamma,EA
            Max, Min, Beta, Gamma, EA = (float(v) for v in vals)
            return dict(logMax=math.log10(Max), logMin=math.log10(Min),
                        Beta=Beta, Gamma=Gamma, EA=EA)
    raise ValueError("未找到主曲线系数块 (Max E* (ksi) ...)")


class MasterCurve:
    """单个 sheet 的主曲线; estar(T_C, f_Hz) -> MPa。"""
    def __init__(self, xlsx_path, sheet, tref_c=TREF_C_DEFAULT):
        wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
        if sheet not in wb.sheetnames:
            raise KeyError(f"sheet 不存在: {sheet} (可选: {wb.sheetnames})")
        self.c = _find_coeffs(wb[sheet])
        self.tref_K = tref_c + 273.15
        self.sheet = sheet

    def log_aT(self, T_C):
        T_K = T_C + 273.15
        return (self.c["EA"] / (LOG10 * R_GAS)) * (1.0 / T_K - 1.0 / self.tref_K)

    def estar_ksi(self, T_C, f_Hz):
        wr = f_Hz * (10.0 ** self.log_aT(T_C))
        x = self.c["Beta"] + self.c["Gamma"] * math.log10(wr)
        logE = self.c["logMin"] + (self.c["logMax"] - self.c["logMin"]) / (1.0 + math.exp(x))
        return 10.0 ** logE

    def estar_MPa(self, T_C, f_Hz):
        return self.estar_ksi(T_C, f_Hz) * KSI_TO_MPA


def estar_MPa(xlsx_path, section, aging, layer, T_C, f_Hz):
    """便捷入口。layer ∈ {'surface','base'}: surface->段自身面层主曲线; base->CG Base。"""
    surf, base = sheet_names(section, aging)
    sheet = surf if layer == "surface" else base
    return MasterCurve(xlsx_path, sheet).estar_MPa(T_C, f_Hz)


if __name__ == "__main__":
    import sys
    xlsx = sys.argv[1] if len(sys.argv) > 1 else "Dynamic_Modulus_Mastercurves.xlsx"
    mc = MasterCurve(xlsx, "N1-1 RH")
    print("=== 自检: N1-1 RH, 对照 sheet 内置平均/拟合值 ===")
    checks = [(20, 10, 1111.4, 7663), (40, 10, 360.5, 2485), (4, 10, 2155.8, 14862),
              (20, 1, 735.5, None), (40, 0.1, 94.3, None)]
    for T, f, ksi_ref, mpa_ref in checks:
        got_ksi = mc.estar_ksi(T, f); got_mpa = mc.estar_MPa(T, f)
        tag = f"(sheet avg≈{ksi_ref} ksi" + (f" ≈{mpa_ref} MPa)" if mpa_ref else ")")
        print(f"  T={T:>3}°C f={f:>4}Hz -> {got_ksi:8.1f} ksi = {got_mpa:8.1f} MPa   {tag}")
    print("\n=== 各段面层 E*(20°C,10Hz) 对照 build_ncat_case (应≈面层E*列) ===")
    for s in ["N1","N2","N5","N8","S5","S6","S13"]:
        try:
            v = estar_MPa(xlsx, s, "RH", "surface", 20, 10)
            print(f"  {s:<4} surface 20°C/10Hz = {v:8.1f} MPa")
        except Exception as e:
            print(f"  {s}: {e}")
    print(f"  CG Base 20°C/10Hz = {estar_MPa(xlsx,'N1','RH','base',20,10):8.1f} MPa (应≈5239)")
    print("\n=== 车辙相关: 各段面层在高温档的 E* (10Hz) ===")
    for T in [25, 35, 45]:
        row = " ".join(f"{estar_MPa(xlsx,s,'RH','surface',T,10):7.0f}" for s in ["N1","N8","S6"])
        print(f"  T={T}°C  N1/N8/S6 = {row} MPa")
