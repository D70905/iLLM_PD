from pathlib import Path
import json
import pandas as pd
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WB = ROOT / "Source_Data" / "SourceData_iLLM-PD_NC_final.xlsx"
TABLES = ROOT / "Source_Data" / "_final_tables_0813"

EXPECTED = [
    "README",
    "Supplementary_source_index",
    "Table1_LTPP_sections",
    "Table2_ablation",
    "Fig2_trajectory",
    "Fig3_LTPP_design_response",
    "Fig4_method_comparison",
    "Fig4_MEPDG_rutting",
    "Fig5_temperature_fatigue",
    "Fig6_OOD_response",
]


def read_xlsx(sheet):
    return pd.read_excel(WB, sheet_name=sheet)


def read_csv(sheet):
    return pd.read_csv(TABLES / f"{sheet}.csv")


def main():
    wb = load_workbook(WB, read_only=True, data_only=True)
    actual = wb.sheetnames
    report = {
        "workbook": str(WB),
        "sheet_order": actual,
        "missing_sheets": [s for s in EXPECTED if s not in actual],
        "extra_sheets": [s for s in actual if s not in EXPECTED],
        "sheets": {},
        "key_checks": {},
        "warnings": [],
    }

    for sheet in EXPECTED:
        if sheet not in actual:
            continue
        df = read_xlsx(sheet)
        src = read_csv(sheet)
        report["sheets"][sheet] = {
            "xlsx_shape": list(df.shape),
            "csv_shape": list(src.shape),
            "columns_match": list(df.columns) == list(src.columns),
            "blank_column_count": int(sum(str(c).startswith("Unnamed") for c in df.columns)),
        }

    if "Fig3_NCAT_boundary" in actual:
        report["warnings"].append("Unexpected Fig3_NCAT_boundary sheet is present.")

    fig2 = read_xlsx("Fig2_trajectory")
    report["key_checks"]["Fig2"] = {
        "section_ids": sorted(fig2["LTPP_section_id"].unique().tolist()),
        "study_sections": sorted(fig2["study_section"].unique().tolist()),
        "n_states": int(len(fig2)),
        "selected_step_3_DSR": float(fig2.loc[fig2["step"] == 3, "DSR"].iloc[0]),
        "step_3_construction_cost_USD_m2": float(fig2.loc[fig2["step"] == 3, "construction_cost_USD_m2"].iloc[0]),
        "final_running_SCR": float(fig2.loc[fig2["step"] == fig2["step"].max(), "running_SCR"].iloc[0]),
        "guard_rejected_steps": [int(x) for x in fig2.loc[fig2["guard_rejected"] == True, "step"].tolist()],
    }

    fig3 = read_xlsx("Fig3_LTPP_design_response")
    report["key_checks"]["Fig3"] = {
        "sections": fig3["study_section"].tolist(),
        "n": int(len(fig3)),
        "all_DSR_1": bool((pd.to_numeric(fig3["final_DSR"], errors="coerce") == 1).all()),
        "all_episode_SCR_1": bool((pd.to_numeric(fig3["episode_SCR"], errors="coerce") == 1).all()),
        "construction_cost_minmax": [
            float(fig3["construction_cost_USD_m2"].min()),
            float(fig3["construction_cost_USD_m2"].max()),
        ],
        "twenty_year_LCC_minmax": [
            float(fig3["twenty_year_LCC_USD_m2"].min()),
            float(fig3["twenty_year_LCC_USD_m2"].max()),
        ],
    }

    by_family = fig3.groupby("pavement_family").agg(
        final_DSR_mean=("final_DSR", "mean"),
        episode_SCR_mean=("episode_SCR", "mean"),
        construction_cost_mean=("construction_cost_USD_m2", "mean"),
        construction_cost_sd=("construction_cost_USD_m2", "std"),
        LCC_mean=("twenty_year_LCC_USD_m2", "mean"),
        LCC_sd=("twenty_year_LCC_USD_m2", "std"),
    )
    report["key_checks"]["Fig3_by_family"] = json.loads(by_family.to_json())

    fig4 = read_xlsx("Fig4_method_comparison")
    report["key_checks"]["Fig4_methods"] = {
        "n": int(len(fig4)),
        "methods": sorted(fig4["method"].unique().tolist()),
        "iLLM_pass": int(((fig4["method"] == "iLLM-PD") & (fig4["JTG_compliant"] == True)).sum()),
        "asbuilt_pass": int(((fig4["method"] == "As-built") & (fig4["JTG_compliant"] == True)).sum()),
        "aashto_pass": int(((fig4["method"] == "AASHTO 1993") & (fig4["JTG_compliant"] == True)).sum()),
    }

    mepdg = read_xlsx("Fig4_MEPDG_rutting")
    report["key_checks"]["Fig4_MEPDG"] = {
        "n": int(len(mepdg)),
        "rutting_pass_count": int((mepdg["MEPDG_total_rutting_mm"] < 19).sum()),
        "max_rutting_mm": float(mepdg["MEPDG_total_rutting_mm"].max()),
        "seeds": sorted([int(x) for x in mepdg["seed"].dropna().unique().tolist()]),
    }

    climate = read_xlsx("Fig5_temperature_fatigue")
    report["key_checks"]["Fig5_temperature_fatigue"] = {
        "n": int(len(climate)),
        "ratio_minmax": [
            float(climate["fixed_over_climate_resolved_fatigue_life_ratio"].min()),
            float(climate["fixed_over_climate_resolved_fatigue_life_ratio"].max()),
        ],
        "MAAT_minmax": [float(climate["MAAT_C"].min()), float(climate["MAAT_C"].max())],
    }

    ood = read_xlsx("Fig6_OOD_response")
    report["key_checks"]["Fig6_OOD"] = {
        "n": int(len(ood)),
        "minimum_B3_min": float(ood["minimum_B3_permanent_deformation_margin"].min()),
        "max_guard_events": int(ood["NumericalGuard_events_total"].max()),
        "max_FEA_escalation_rate": float(ood["FEA_escalation_rate"].max()),
        "no_spurious_final_DSR_gt_1": bool((pd.to_numeric(ood["final_DSR"], errors="coerce") <= 1).all()),
    }

    table2 = read_xlsx("Table2_ablation")
    report["key_checks"]["Table2"] = {
        "n": int(len(table2)),
        "variants": sorted(table2["variant"].unique().tolist()),
        "families": sorted(table2["pavement_family"].unique().tolist()),
    }

    report["passed"] = (
        not report["missing_sheets"]
        and not report["extra_sheets"]
        and all(s["columns_match"] for s in report["sheets"].values())
        and all(s["blank_column_count"] == 0 for s in report["sheets"].values())
        and not report["warnings"]
    )
    print(json.dumps(report, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
